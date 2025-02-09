import streamlit as st
#from emotion_affirm.py import analyze_mood, generate_affirmation 
import random
import os
from openpyxl import Workbook, load_workbook

# model.py
import joblib
import pandas as pd
import spacy
import schedule
import time
import smtplib
from email.mime.text import MIMEText
from sklearn.feature_extraction.text import TfidfVectorizer

max_mood = ""
# Load the model
def load_model():
    model = joblib.load('simbolo/Project/random_forest_model (1).pkl')  # Replace with the actual path to your model
    return model

def load_vectorizer():
    vectorizer = joblib.load('simbolo/Project/vectorizer.pkl')
    return vectorizer


model = load_model()
vectorizer = load_vectorizer()

affirmations = pd.read_csv('simbolo/Project/affirmation new.csv')  # Affirmation Dataset
nlp = spacy.load('en_core_web_sm')



def preprocess_text_spacy(text):
    # Convert text to lowercase and process with spaCy
    doc = nlp(text.lower())
    # Remove punctuation and stopwords, and tokenize
    tokens = [token.text for token in doc if not token.is_stop and not token.is_punct]
    return ' '.join(tokens)

def analyze_mood(user_input):
    cleaned_input = preprocess_text_spacy(user_input)
    input_vector = vectorizer.transform([cleaned_input]).toarray()
    mood_probabilities = model.predict_proba(input_vector)[0]
    mood_classes = model.classes_
    mood_analysis = {mood_classes[i]: mood_probabilities[i] for i in range(len(mood_classes))}
    return mood_analysis
def generate_affirmation(mood_analysis):
    predicted_mood = max(mood_analysis, key=mood_analysis.get)
    filtered_affirmations = affirmations[affirmations['Mood'] == predicted_mood]['Affirmation'].tolist()
    return random.choice(filtered_affirmations)
# Save User Email and Affirmation
# Step 10: Save User Email and Affirmation

# Step 10: Save User Email and Affirmation
file_path = 'user_affirmations.xlsx'

def save_email_and_affirmation(email, affirmation):

    # Create a new workbook if the file doesn't exist
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Email", "Affirmation"])
        workbook.save(file_path)

    # Add the email and affirmation to the file
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet.append(["Email", "Affirmation"])
    sheet.append([email, affirmation])
    workbook.save(file_path)

# Send Affirmation Email
def send_email(email, affirmation):
    sender_email = "posivibescorner@gmail.com"  # Replace with your email address
    sender_password = "tqvfufqvveqgziki"  # Replace with your Gmail app password
    subject = "Your Morning Affirmation 🌞"
    body = f"Good morning!\n\nAre you still feeling {max_mood}?If so,please repeat this:\n\n{affirmation}\n\nHave a wonderful day !\nPlease spread peace and positivity\n\nWith love and peace,\nPosiVibes Team"

    msg = MIMEText(body, "plain", "utf-8")
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = email

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, email, msg.as_string())
    except Exception as e:
        print("Error sending email:", e)
 #Step 12: Schedule Email Sending


def send_scheduled_emails():
    # Check if the file exists
    if not os.path.exists(file_path):
        print("No emails to send.")
        return

    # If the file exists, load the existing workbook
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # If no rows exist (except header), print and return
    if sheet.max_row == 1:
        print("No email data to send.")
        return

    # Iterate over the rows starting from row 2 (skip header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
      
        email, affirmation = row
        send_email(email, affirmation)


    # Clear the data after sending the emails (keeping header)
    sheet.delete_rows(2, sheet.max_row)
    workbook.save(file_path)
    print("Emails sent and data cleared.")

# Title
st.title("Posivibes Affirmation Generator")

# Subtitle
st.subheader("Start your day with positivity and motivation!")

# Description
st.write(
    "Feeling down or in need of encouragement? Generate a powerful affirmation "
    "to uplift your spirit and boost your confidence."
)

user_input = st.text_input("How are you feeling today?", placeholder="Type your thoughts or feelings here...")
email = st.text_input('Please enter your email address:')


if st.button("Generate Affirmation"):
    if user_input and email:
        # Analyze mood and generate affirmation
        mood_analysis = analyze_mood(user_input)
        st.write("Mood Analysis:") 
        max_mood = max(mood_analysis, key=mood_analysis.get)

        for mood, prob in mood_analysis.items():
            st.write(f"{mood}: {prob*100:.2f}%")

        affirmation = generate_affirmation(mood_analysis)
        st.write("\nHere’s your affirmation for today:")
        st.success(affirmation)
        
        # Save email and affirmation (optional)
        save_email_and_affirmation(email, affirmation)
        st.write(f"Your affirmation has been saved and will be sent to {email} tomorrow morning.") 
        send_scheduled_emails()

     

    else:
        st.warning('Please enter both your mood and email address')
